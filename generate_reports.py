import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def load_json(filepath):
    if not os.path.exists(filepath):
        print(f"Warning: {filepath} not found.")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return []

def create_excel_report(data, filename, report_title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    pass_font = Font(color="008000", bold=True)
    fail_font = Font(color="FF0000", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Report Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = report_title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = center_align

    # Headers
    headers = ["Test ID", "Category", "Test Name", "Type", "Status", "Duration (ms)", "Error Message"]
    ws.append([]) # Empty row for spacing
    ws.append(headers)

    header_row = ws[3]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    total, passed, failed = 0, 0, 0
    for item in data:
        status = item.get("status", "FAIL")
        total += 1
        if status == "PASS":
            passed += 1
        else:
            failed += 1

        row = [
            item.get("id", ""),
            item.get("category", ""),
            item.get("name", ""),
            item.get("type", ""),
            status,
            item.get("duration", 0),
            item.get("error", "")
        ]
        ws.append(row)
        
        current_row = ws[ws.max_row]
        for idx, cell in enumerate(current_row):
            cell.border = thin_border
            if idx == 4: # Status column
                cell.font = pass_font if status == "PASS" else fail_font
                cell.alignment = center_align
            elif idx in (0, 3, 5):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    column_widths = [12, 25, 60, 15, 12, 15, 50]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = column_width

    # Summary Sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    summary_data = [
        ("Report Title", report_title),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Pass Rate", f"{(passed/total)*100:.2f}%" if total > 0 else "0.00%")
    ]
    
    for row in summary_data:
        ws_summary.append(row)
        
    for r in ws_summary.iter_rows(min_row=1, max_row=len(summary_data)):
        r[0].font = Font(bold=True)
        r[0].border = thin_border
        r[1].border = thin_border

    # Ensure output directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"Generated report: {filename} ({total} tests)")


def generate_github_summary(web_data, mobile_data):
    """Write a comprehensive per-test Step Summary and GitHub annotations."""
    summary_env = os.environ.get("GITHUB_STEP_SUMMARY")

    # ── Stats ───────────────────────────────────────────────────
    web_total   = len(web_data)
    web_passed  = sum(1 for d in web_data if d.get("status") == "PASS")
    web_failed  = sum(1 for d in web_data if d.get("status") == "FAIL")
    web_skipped = sum(1 for d in web_data if d.get("status") == "SKIP")
    web_pct     = (web_passed / web_total * 100) if web_total > 0 else 0

    mob_total   = len(mobile_data)
    mob_passed  = sum(1 for d in mobile_data if d.get("status") == "PASS")
    mob_failed  = sum(1 for d in mobile_data if d.get("status") == "FAIL")
    mob_skipped = sum(1 for d in mobile_data if d.get("status") == "SKIP")
    mob_pct     = (mob_passed / mob_total * 100) if mob_total > 0 else 0

    grand_total  = web_total  + mob_total
    grand_passed = web_passed + mob_passed
    grand_failed = web_failed + mob_failed

    md = []

    # ── Header ──────────────────────────────────────────────────
    md.append("# 🍱 FoodBridge — Complete Test Results\n")
    md.append(f"> **{grand_passed}/{grand_total}** tests passed  |  "
              f"Run: `{datetime.now().strftime('%Y-%m-%d %H:%M UTC')}`\n")

    # ── Overall summary table ────────────────────────────────────
    md.append("## 📊 Suite Overview\n")
    md.append("| Suite | Total | ✅ Passed | ❌ Failed | ⏭ Skipped | Pass Rate |")
    md.append("|:---|:---:|:---:|:---:|:---:|:---:|")
    if web_total > 0:
        md.append(f"| 🌐 **Selenium Web E2E** | {web_total} | {web_passed} | {web_failed} | {web_skipped} | **{web_pct:.1f}%** |")
    else:
        md.append("| 🌐 **Selenium Web E2E** | — | — | — | — | *no data* |")
    if mob_total > 0:
        md.append(f"| 📱 **Appium Mobile E2E** | {mob_total} | {mob_passed} | {mob_failed} | {mob_skipped} | **{mob_pct:.1f}%** |")
    else:
        md.append("| 📱 **Appium Mobile E2E** | — | — | — | — | *no data* |")
    md.append(f"| **TOTAL** | **{grand_total}** | **{grand_passed}** | **{grand_failed}** | — | **{(grand_passed/grand_total*100):.1f}%** |" if grand_total > 0 else "")
    md.append("\n---\n")

    # ── Helper: write all tests in a suite grouped by category ──
    def write_suite_tests(title, icon, data, suite_label):
        if not data:
            md.append(f"## {icon} {title}\n")
            md.append("*No test data recorded for this suite.*\n")
            return

        md.append(f"## {icon} {title}\n")

        # Group by category
        cats = {}
        for t in data:
            cat = t.get("category", "General")
            cats.setdefault(cat, []).append(t)

        for cat, tests in cats.items():
            cat_pass = sum(1 for t in tests if t.get("status") == "PASS")
            cat_fail = sum(1 for t in tests if t.get("status") == "FAIL")
            cat_icon = "🟢" if cat_fail == 0 else "🔴"
            cat_rate = (cat_pass / len(tests) * 100) if tests else 0

            md.append(f"### {cat_icon} {cat} &nbsp; `{cat_pass}/{len(tests)} passed` ({cat_rate:.0f}%)\n")
            md.append("| Status | Test ID | Test Name | Duration |")
            md.append("|:---:|:---|:---|:---:|")

            for t in tests:
                status   = t.get("status", "FAIL")
                s_icon   = "✅" if status == "PASS" else ("⏭" if status == "SKIP" else "❌")
                tid      = t.get("id", "—")
                name     = t.get("name", "Unknown test")
                dur_ms   = t.get("duration", 0)
                dur_str  = f"{dur_ms}ms" if dur_ms < 1000 else f"{dur_ms/1000:.1f}s"
                err      = t.get("error", "")

                if status == "FAIL" and err:
                    # Show error inline in a details block
                    md.append(f"| {s_icon} | `{tid}` | **{name}** ❗ `{err[:120]}` | {dur_str} |")
                else:
                    md.append(f"| {s_icon} | `{tid}` | {name} | {dur_str} |")

            md.append("")  # blank line between categories

        md.append("\n---\n")

    # ── Selenium Web tests (all categories including Vulnerability) ─
    write_suite_tests(
        "Selenium Web E2E — All Test Cases",
        "🌐", web_data, "Selenium Web"
    )

    # ── Appium Mobile tests ──────────────────────────────────────
    write_suite_tests(
        "Appium Mobile E2E — All Test Cases",
        "📱", mobile_data, "Appium Mobile"
    )

    # ── Write to GITHUB_STEP_SUMMARY ────────────────────────────
    if summary_env:
        try:
            with open(summary_env, "a", encoding="utf-8") as f:
                f.write("\n".join(md) + "\n")
            print("✅ GitHub Step Summary written successfully.")
        except Exception as e:
            print(f"❌ Error writing Step Summary: {e}")
    else:
        print("ℹ️  GITHUB_STEP_SUMMARY not set (not running in GitHub Actions)")

    # ── GitHub Annotations (::notice:: / ::error::) ──────────────
    # These appear in the "Annotations" panel of the Actions run
    for t in web_data:
        status = t.get("status", "FAIL")
        name   = t.get("name", "Unknown")
        tid    = t.get("id", "")
        label  = f"{tid} — {name}" if tid else name
        if status == "PASS":
            print(f"::notice title=✅ {label}::PASS")
        elif status == "SKIP":
            print(f"::notice title=⏭ {label}::SKIPPED")
        else:
            err = t.get("error", "Test failed")
            print(f"::error title=❌ {label}::{err[:200]}")

    for t in mobile_data:
        status = t.get("status", "FAIL")
        name   = t.get("name", "Unknown")
        tid    = t.get("id", "")
        label  = f"{tid} — {name}" if tid else name
        if status == "PASS":
            print(f"::notice title=✅ {label}::PASS")
        elif status == "SKIP":
            print(f"::notice title=⏭ {label}::SKIPPED")
        else:
            err = t.get("error", "Test failed")
            print(f"::error title=❌ {label}::{err[:200]}")


def main():
    web_file = os.path.join(".web_results.json")
    mobile_file = os.path.join("appium-python-tests", ".pytest_results.json")
    
    web_data = load_json(web_file)
    mobile_data = load_json(mobile_file)
    
    # 1. Selenium Report (Exclude Vulnerability)
    selenium_tests = [d for d in web_data if "Vulnerability" not in d.get("category", "")]
    if selenium_tests:
        create_excel_report(
            selenium_tests, 
            os.path.join("reports", "Selenium_Web_Test_Report.xlsx"), 
            "Selenium Web Functional & UI Test Report"
        )
        
    # 2. Appium Report (Exclude Vulnerability)
    appium_tests = [d for d in mobile_data if "Vulnerability" not in d.get("category", "")]
    if appium_tests:
        create_excel_report(
            appium_tests, 
            os.path.join("reports", "Appium_Mobile_Test_Report.xlsx"), 
            "Appium Mobile Functional & E2E Test Report"
        )
        
    # 3. Vulnerability Report (Only Vulnerability from Web and Mobile)
    vuln_tests = [d for d in web_data if "Vulnerability" in d.get("category", "")] + \
                 [d for d in mobile_data if "Vulnerability" in d.get("category", "")]
    if vuln_tests:
        create_excel_report(
            vuln_tests, 
            os.path.join("reports", "Security_Vulnerability_Test_Report.xlsx"), 
            "Comprehensive Security & Vulnerability Test Report"
        )

    # Generate GitHub Actions summary
    generate_github_summary(web_data, mobile_data)


if __name__ == "__main__":
    main()
