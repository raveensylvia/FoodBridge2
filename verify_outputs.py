#!/usr/bin/env python3
import os, json, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace') if hasattr(sys.stdout, 'reconfigure') else None
from openpyxl import load_workbook

root = r'c:\Users\ravee\Downloads\food_bridge2'

print("=== JSON FILES ===")
for p in ['.web_results.json', 'appium-python-tests/.pytest_results.json']:
    fullp = os.path.join(root, p)
    with open(fullp, encoding='utf-8') as f:
        d = json.load(f)
    passed = sum(1 for t in d if t["status"] == "PASS")
    failed = sum(1 for t in d if t["status"] == "FAIL")
    print(f"  {p}: {len(d)} tests | PASS={passed} FAIL={failed}")

print()
print("=== EXCEL REPORTS in tests/ ===")
tests_dir = os.path.join(root, 'tests')
for fname in sorted(os.listdir(tests_dir)):
    if not fname.endswith('.xlsx'):
        continue
    wb = load_workbook(os.path.join(tests_dir, fname))
    ws_sum = wb['Summary']
    rows = list(ws_sum.iter_rows(min_row=2, max_row=2, values_only=True))
    if rows:
        row = rows[0]
        print(f"  {fname}")
        print(f"    Suite  : {row[0]}")
        print(f"    Total  : {row[1]} | Passed: {row[2]} | Failed: {row[3]} | Rate: {row[4]}%")
        print(f"    Sheets : {wb.sheetnames}")

print()
print("=== JUNIT XMLs (test-results/) ===")
xml_dir = os.path.join(root, 'test-results')
if os.path.exists(xml_dir):
    for fname in sorted(os.listdir(xml_dir)):
        if fname.endswith('.xml'):
            size = os.path.getsize(os.path.join(xml_dir, fname))
            print(f"  {fname}: {size:,} bytes")

print()
print("=== REPORTS (reports/) ===")
rep_dir = os.path.join(root, 'reports')
if os.path.exists(rep_dir):
    for fname in sorted(os.listdir(rep_dir)):
        if fname.endswith('.xlsx'):
            size = os.path.getsize(os.path.join(rep_dir, fname))
            print(f"  {fname}: {size:,} bytes")

print()
print("ALL CHECKS DONE.")
